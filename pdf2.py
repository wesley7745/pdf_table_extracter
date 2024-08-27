import time
import streamlit as st
import numpy as np
import pandas as pd
import io
import openpyxl
import pdfplumber


def search_keyword_in_pdf(pdf_path, keyword):
    pdf = pdfplumber.open(pdf_path)
    total_pages = len(pdf.pages)

    pages_with_keyword = []

    for page_num in range(4, total_pages):  # 從第五頁開始計數
        page = pdf.pages[page_num]
        text = page.extract_text()
        if text and keyword.lower() in text.lower():
            pages_with_keyword.append(page_num + 1 - 4)  # 調整後的頁碼，從1開始計數

    pdf.close()
    return pages_with_keyword

def extract_tables_from_pdf(pdf_path, pages):
    pdf = pdfplumber.open(pdf_path)
    page_tables = []

    for page_num in pages:
        page = pdf.pages[page_num + 3]  # 頁碼從0開始，跳過前四頁
        tables = page.extract_tables()
        if tables:
            page_tables.append((page_num, tables))
    
    pdf.close()
    return page_tables

def create_excel_buffer(page_tables):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        for page_num, tables in page_tables:
            for idx, table in enumerate(tables):
                df = pd.DataFrame(table[1:], columns=table[0])
                df_transposed = df.T  # 將DataFrame轉置
                sheet_name = f"Page_{page_num}_Table_{idx + 1}"
                df_transposed.to_excel(writer, sheet_name=sheet_name, header=False)  # 不保存原來的header
                st.write(f"頁碼 {page_num} 的表格 {idx + 1} 已保存到 Excel 文件的工作表 {sheet_name}")
    output.seek(0)
    return output

def filter_sheets_by_content(file_path, keywords):
    # Open the original workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Create a new workbook to save the filtered sheets
    new_wb = openpyxl.Workbook()
    # Remove the default sheet created with the new workbook if it's empty
    if len(new_wb.sheetnames) == 1 and new_wb.sheetnames[0] == 'Sheet':
        new_wb.remove(new_wb.active)
    
    # Iterate over all sheets in the workbook
    for sheet_name in wb.sheetnames:
        source_sheet = wb[sheet_name]
        # Check if any cell in the sheet contains the keywords
        contains_keyword = False
        for row in source_sheet.iter_rows(values_only=True):
            if any(keyword in str(cell) for cell in row if cell is not None for keyword in keywords):
                contains_keyword = True
                break
        
        if contains_keyword:
            new_sheet = new_wb.create_sheet(title=sheet_name)
            for row in source_sheet.iter_rows(values_only=True):
                new_sheet.append(row)
    
    # Save the new workbook
    new_file_path = file_path.replace('.xlsx', '_filtered.xlsx')
    new_wb.save(new_file_path)
    return new_file_path


def combine_and_deduplicate_excel(files):
    combined_df = pd.DataFrame()
    for file in files:
        xls = pd.ExcelFile(file)
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            combined_df = pd.concat([combined_df, df], ignore_index=True)
    
    # 篩選掉包含 "截至" 或 "月底" 的行
    keywords = ['截至', '月底']
    combined_df = combined_df[~combined_df.apply(lambda row: row.astype(str).str.contains('|'.join(keywords)).any(), axis=1)]
    
    # 去重
    combined_df.drop_duplicates(inplace=True)
    
    # 處理未命名列
    combined_df.columns = [None if 'Unnamed' in str(col) else col for col in combined_df.columns]
    
    # 將結果寫入 Excel
    output = io.BytesIO()
    combined_df.to_excel(output, index=False, header=True)
    output.seek(0)
    return output


st.set_page_config(
    page_title="檔案表格提取小助手",
    page_icon=":hammer:",
    layout="centered",
    initial_sidebar_state="expanded",
)

st.title("檔案表格提取小助手")
st.write("可以透過這個小助手，提取年檔案中的表格，但是要提取的表格必須有框線，否則會提取不到。")

tab1, tab2, tab3 = st.tabs(["年報關鍵字", "檔案清理", "報表結合"])

with tab1:
    st.title('年報關鍵字')
    
    # 文件上傳
    uploaded_files = st.sidebar.file_uploader("**上傳文件**(可以一次上傳多個檔案)", type='pdf', accept_multiple_files=True)

    keyword = st.sidebar.text_input("**請輸入關鍵字**", " ")
    st.sidebar.markdown(":warning: 請確保關鍵字的正確性，否則可能無法提取到表格，建議不要用複製貼上的方式輸入關鍵字。")

    name = st.sidebar.text_input("**輸入文件名稱**", "extracted_tables.xlsx")
    st.sidebar.markdown("EX: extracted_tables.xlsx")

    if st.sidebar.button("提取表格"):
        if uploaded_files and keyword:
            all_page_tables = []
            for uploaded_file in uploaded_files:
                _path = uploaded_file.name
                with open(_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())

                pages_with_keyword = search_keyword_in_pdf(_path, keyword)
                st.write(f"關鍵字 '{keyword}' 出現在以下頁面: {pages_with_keyword}")

                page_tables = extract_tables_from_pdf(_path, pages_with_keyword)
                all_page_tables.extend(page_tables)

            excel_buffer = create_excel_buffer(all_page_tables)
            st.success("表格已提取完畢！")

            st.download_button(
                label="下載Excel文件",
                data=excel_buffer,
                file_name=name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # 創建 DataFrame
    df = pd.DataFrame({
        '公司': ['華邦電', '台積電', '旺宏', '南亞科', '晶豪科', '新唐', '力積電', '鈺創'],
        '關鍵字': ['從業員工', '人力結構', '從業員工', '從業員工', '從業員工', '從業員工', '從業員工', '員工人數']
    })
    st.table(df.reset_index(drop=True))

with tab2:
    st.title('檔案清理')

    # 文件上傳
    excel_files = st.file_uploader("**上傳Excel文件**(可以一次上傳多個檔案)", type='xlsx', accept_multiple_files=True, key="clean")

    keyword1 = st.text_input("**請輸入關鍵字**(只會保留包含此關鍵字的sheet)", " ")
    keywords = keyword1.split(',')  # 假設用戶用逗號分隔多個關鍵字

    if st.button("清理檔案"):
        if excel_files and keyword1:
            for uploaded_file in excel_files:
                with open(uploaded_file.name, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                filtered_file_path = filter_sheets_by_content(uploaded_file.name, keywords)

                with open(filtered_file_path, "rb") as f:
                    st.download_button(
                        label=f"下載清理後的 {uploaded_file.name}",
                        data=f,
                        file_name=f"{uploaded_file.name.replace('.xlsx', '_filtered.xlsx')}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            st.success("檔案已清理，只保留包含關鍵字的工作表！")

with tab3:
    st.title('報表結合')

    # 文件上傳
    excel_files = st.file_uploader("**上傳Excel文件**(可以一次上傳多個檔案)", type='xlsx', accept_multiple_files=True)

    fname = st.text_input("**輸入文件名稱**", " ")

    if st.button("合併並去重"):
        if excel_files:
            combined_excel = combine_and_deduplicate_excel(excel_files)
            st.success("報表已合併並去重！")

            st.download_button(
                label="下載合併並去重後的Excel文件",
                data=combined_excel,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
